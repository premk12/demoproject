import openpyxl


def getRowcount(file,sheetname):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook.active
    return (sheet.max_row)

def getRowcount(file,sheetname):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook.active
    return (sheet.min_column)

def readData(file, sheetname, rownum, colnum):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook.active
    inputvalue = sheet.cell(row=rownum, column=colnum).value
    return inputvalue


def writeData(file, sheetname, rownum, colnum, data):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook.active
    sheet.cell(row=rownum, column=colnum).value = data
    workbook.save(file)