import openpyxl

filename = "c:\\UserIDPass.xlsx"
# Load/ open excel file
workbook= openpyxl.load_workbook(filename)

# Locate /get perticular sheet
###sheet = workbook.get_sheet_by_name("UserID")  ## older version
sheet = workbook["UserID"]

# Read data from sheet
print(sheet.cell(row=1, column=1).value)

print(sheet.cell(row=7, column=2).value)

# Assignment : Read all values from excel

## Write operation
temp = sheet.cell(row=8, column=1).value
sheet.cell(row=8, column=1).value = "Pratik"
sheet.cell(row=8, column=2).value = "PratikGreaterKing"

print(sheet.cell(row=8, column=1).value +  "  "  + sheet.cell(row=8, column=2).value)

workbook.save(filename)


