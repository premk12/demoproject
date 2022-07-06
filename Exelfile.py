import openpyxl
file="C:\\Users\\Prem\\Downloads\\Exeldata.xlsx"

workbook=openpyxl.load_workbook(file)
sheet=workbook.active   ## for multiple sheet

rows=sheet.max_row
column=sheet.max_column

print(rows)
print(column)


####read all the rows and column

for r in range(1,rows+1):
    for c in range(1,column+1):
     print(sheet.cell(r,c).value, end="     ")

    print()


### / insert data from pythod to exel

# workbook=openpyxl.load_workbook(file)
#
# sheet=workbook.active
#
# sheet.cell(1,3).value=123
# sheet.cell(1,2).value="prem"
# sheet.cell(2,1).value=124
# sheet.cell(2,2).value="amul"
# sheet.cell(3,1).value=125
# sheet.cell(3,2).value="punam"
#
# workbook.save(file)
# #
#
#







