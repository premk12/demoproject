import openpyxl
import time
from selenium import webdriver
from selenium.webdriver.common.by import By
from openpyxl import load_workbook
from selenium import webdriver
from selenium.webdriver.support.select import Select

driver=webdriver.Chrome()

driver.implicitly_wait(10)
driver.maximize_window()

driver.get("https://practicetestautomation.com/practice-test-login/")

import ExelUtiles



file="C:\\Users\\Prem\\Downloads\\Bookprem.xlsx"

wb = load_workbook(file)
sheet=wb.active
row_count = sheet.max_row
col_count = sheet.max_column
for r in range(2,row_count+1):
    username=ExelUtiles.readData(file, "Sheet1", r, 2)
    password = ExelUtiles.readData(file, "Sheet1", r, 3)
    login=driver.find_element(By.ID,'username').send_keys(username)
    password=driver.find_element(By.ID,'password').send_keys(password)

    click=driver.find_element(By.ID,'submit').click()

    pagetitle = driver.title
    print(pagetitle)
    if pagetitle == "Logged In Successfully | Practice Test Automation":
        ExelUtiles.writeData(file, "Sheet1", r, 4, "Passed")
    else:
        ExelUtiles.writeData(file, "Sheet1", r, 4, "Failed")










